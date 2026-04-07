#!/usr/bin/env python3
"""
BTG TMT — Telecom Pricing Monitor
===================================
Scrapes Brazilian carrier websites for entry-level mobile plan data, detects
changes against the stored plans.json, creates GitHub Issues for alerts
(GitHub emails you automatically), and logs every change to a persistent
XLSX changelog.

Notifications work via GitHub Issues — no email credentials needed.
GitHub sends an email to the repo owner whenever an issue is opened.

Usage
-----
  Normal run (scrape + diff + issue + write):
    python scraper/scrape.py

  Dry run (scrape + diff, no writes, no issue created):
    python scraper/scrape.py --dry-run

  Force notification even when nothing changed (test the flow):
    python scraper/scrape.py --force-notify

Environment variables (set as GitHub Actions Secrets)
-----------------------------------------------------
  GITHUB_TOKEN     automatically provided by GitHub Actions (no setup needed)
  GITHUB_REPO      set automatically by Actions as: owner/repo-name
"""

import argparse
import json
import logging
import os
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Bootstrap ─────────────────────────────────────────────────────────────────

# Load .env when running locally (ignored in Actions where env vars come from secrets)
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass

REPO_ROOT      = Path(__file__).parent.parent
PLANS_JSON     = REPO_ROOT / "data" / "plans.json"
CHANGELOG_XLSX = REPO_ROOT / "data" / "changelog.xlsx"

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO  = os.environ.get("GITHUB_REPO", "")   # e.g. "brunof/telecom-tracker"

# BRT = UTC-3 (no DST in Brazil since 2019)
UTC_OFFSET_HOURS = -3

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("monitor")


# ── Time helpers ──────────────────────────────────────────────────────────────

def now_brt() -> datetime:
    """Return current time adjusted to BRT (UTC-3)."""
    from datetime import timezone, timedelta
    utc_now = datetime.now(tz=timezone.utc)
    brt_now = utc_now.astimezone(timezone(timedelta(hours=UTC_OFFSET_HOURS)))
    return brt_now


def today_brt() -> str:
    return now_brt().strftime("%Y-%m-%d")


# ── Page fetcher (Playwright with stealth) ────────────────────────────────────

def fetch_page_text(url: str, wait_seconds: int = 8) -> Optional[str]:
    """
    Fetch a URL using Playwright (headless Chromium) and return the full
    rendered text content of the page. Returns None if the page cannot be
    reached or times out.

    Uses realistic browser headers and disables webdriver detection to
    reduce the chance of being blocked by anti-bot systems.
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        log.error("Playwright not installed. Run: pip install playwright && playwright install chromium")
        return None

    log.info(f"Fetching: {url}")
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                    "--disable-setuid-sandbox",
                ],
            )
            ctx = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale="pt-BR",
                timezone_id="America/Sao_Paulo",
                viewport={"width": 1280, "height": 900},
                java_script_enabled=True,
            )
            # Remove webdriver flag
            ctx.add_init_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            page.wait_for_timeout(wait_seconds * 1000)

            # Dismiss cookie banners (common patterns on BR carrier sites)
            for selector in [
                "button:has-text('Aceitar')",
                "button:has-text('Aceitar todos')",
                "button:has-text('OK')",
                "#onetrust-accept-btn-handler",
                ".cookie-accept",
            ]:
                try:
                    page.click(selector, timeout=2000)
                    page.wait_for_timeout(1000)
                except Exception:
                    pass

            text = page.inner_text("body")
            browser.close()
            return text

    except Exception as exc:
        log.warning(f"Failed to fetch {url}: {exc}")
        return None


# ── Price & GB extraction helpers ─────────────────────────────────────────────

# Matches: R$30, R$ 30, R$59,90, R$ 119,90 — returns float
_PRICE_RE = re.compile(r"R\$\s*(\d{1,4}(?:[.,]\d{2})?)", re.IGNORECASE)

# Matches: 10GB, 10 GB, 10gb
_GB_RE = re.compile(r"(\d{1,3})\s*[Gg][Bb]")


def _to_float(raw: str) -> float:
    """Convert Brazilian price string to float: '59,90' → 59.90"""
    return float(raw.replace(",", "."))


def extract_all_prices(text: str) -> list[float]:
    """Return all R$ prices found in page text, sorted ascending."""
    matches = _PRICE_RE.findall(text)
    prices = []
    for m in matches:
        try:
            prices.append(_to_float(m))
        except ValueError:
            pass
    return sorted(set(prices))


def extract_all_gb(text: str) -> list[int]:
    """Return all GB values found in page text, sorted ascending."""
    matches = _GB_RE.findall(text)
    gb_vals = []
    for m in matches:
        try:
            gb_vals.append(int(m))
        except ValueError:
            pass
    return sorted(set(gb_vals))


def find_price_near_text(text: str, keyword: str, window: int = 300) -> Optional[float]:
    """
    Find a price in the text within `window` characters of a keyword.
    Returns the first price found in that window, or None.
    """
    idx = text.lower().find(keyword.lower())
    if idx == -1:
        return None
    snippet = text[max(0, idx - window // 2): idx + window // 2]
    prices = extract_all_prices(snippet)
    return prices[0] if prices else None


def find_gb_near_text(text: str, keyword: str, window: int = 300) -> Optional[int]:
    """Find a GB value near a keyword. Returns first match or None."""
    idx = text.lower().find(keyword.lower())
    if idx == -1:
        return None
    snippet = text[max(0, idx - window // 2): idx + window // 2]
    gbs = extract_all_gb(snippet)
    return gbs[0] if gbs else None


# ── Carrier-specific parsers ──────────────────────────────────────────────────
# Each parser returns a dict with keys:
#   price (float|None), plan_gb (int|None), confidence ("high"|"low")
#
# confidence="high"  → primary keyword matched; value is trustworthy.
#                      detect_changes will use this value and may update plans.json.
# confidence="low"   → fell back to range scan or secondary keyword; value is a guess.
#                      detect_changes will skip the update and flag for manual review.
#
# Returns None if the page text is empty/invalid.

def parse_vivo_prepaid(text: str) -> Optional[dict]:
    """Vivo pre-paid: site consistently 403. No reliable keyword. Always low."""
    if not text or len(text) < 100:
        return None
    price = find_price_near_text(text, "Turbo", window=400)
    if price is None:
        prices = [p for p in extract_all_prices(text) if 20 <= p <= 60]
        price = prices[0] if prices else None
    gb = find_gb_near_text(text, "Turbo", window=400)
    if gb is None:
        gb = find_gb_near_text(text, "30 dias", window=500)
    return {"price": price, "plan_gb": gb, "confidence": "low"}


def parse_vivo_controle(text: str) -> Optional[dict]:
    """Vivo controle: no reliable primary keyword yet. Always low."""
    if not text or len(text) < 100:
        return None
    price = find_price_near_text(text, "Controle", window=500)
    if price is None:
        prices = [p for p in extract_all_prices(text) if 40 <= p <= 120]
        price = prices[0] if prices else None
    gb = find_gb_near_text(text, "21", window=200)
    if gb is None:
        gbs = [g for g in extract_all_gb(text) if 15 <= g <= 50]
        gb = gbs[0] if gbs else None
    return {"price": price, "plan_gb": gb, "confidence": "low"}


def parse_vivo_postpaid(text: str) -> Optional[dict]:
    """Vivo post-paid: no reliable primary keyword yet. Always low."""
    if not text or len(text) < 100:
        return None
    price = find_price_near_text(text, "Pós", window=500)
    if price is None:
        prices = [p for p in extract_all_prices(text) if 100 <= p <= 250]
        price = prices[0] if prices else None
    gbs = [g for g in extract_all_gb(text) if 30 <= g <= 200]
    gb = gbs[0] if gbs else None
    return {"price": price, "plan_gb": gb, "confidence": "low"}


def parse_tim_prepaid(text: str) -> Optional[dict]:
    if not text or len(text) < 100:
        return None
    confidence = "high"
    # "R$ 30 por 30 dias" — price sits immediately before "por 30 dias"
    price = find_price_near_text(text, "por 30 dias", window=60)
    if price is None:
        confidence = "low"
        price = find_price_near_text(text, "XIP", window=400)
        if price is None:
            prices = [p for p in extract_all_prices(text) if 20 <= p <= 60]
            price = prices[0] if prices else None
    # "12GB da oferta + 4GB exclusivo para redes sociais"
    gb = find_gb_near_text(text, "da oferta", window=60)
    if gb is None:
        confidence = "low"
        gb = find_gb_near_text(text, "XIP", window=400)
        if gb is None:
            gbs = [g for g in extract_all_gb(text) if 6 <= g <= 20]
            gb = gbs[0] if gbs else None
    return {"price": price, "plan_gb": gb, "confidence": confidence}


def parse_tim_controle(text: str) -> Optional[dict]:
    if not text or len(text) < 100:
        return None
    confidence = "high"
    # Card (show state): "R$ 49,99/mês" — price immediately before "/mês"
    price = find_price_near_text(text, "/mês", window=60)
    if price is None:
        confidence = "low"
        prices = [p for p in extract_all_prices(text) if 40 <= p <= 120]
        price = prices[0] if prices else None
    # Card (show state): <h2>45GB</h2> — first plausible GB on page
    gbs = [g for g in extract_all_gb(text) if 20 <= g <= 80]
    gb = gbs[0] if gbs else None
    if gb is None:
        confidence = "low"
    return {"price": price, "plan_gb": gb, "confidence": confidence}


def parse_tim_postpaid(text: str) -> Optional[dict]:
    if not text or len(text) < 100:
        return None
    confidence = "high"
    # "R$ 119,99/mês na fatura" — price immediately before "na fatura"
    price = find_price_near_text(text, "na fatura", window=80)
    if price is None:
        confidence = "low"
        prices = [p for p in extract_all_prices(text) if 100 <= p <= 300]
        price = prices[0] if prices else None
    # <h2>Até 70GB</h2> — GB immediately after "Até"
    gb = find_gb_near_text(text, "Até", window=60)
    if gb is None:
        confidence = "low"
        gbs = [g for g in extract_all_gb(text) if 40 <= g <= 150]
        gb = gbs[0] if gbs else None
    return {"price": price, "plan_gb": gb, "confidence": confidence}


def parse_claro_prepaid(text: str) -> Optional[dict]:
    if not text or len(text) < 100:
        return None
    confidence = "high"
    # "R$30/ 30dias" — price immediately before "30dias"
    price = find_price_near_text(text, "30dias", window=60)
    if price is None:
        price = find_price_near_text(text, "/ 30dias", window=60)
    if price is None:
        confidence = "low"
        price = find_price_near_text(text, "Prezão", window=400)
        if price is None:
            prices = [p for p in extract_all_prices(text) if 20 <= p <= 60]
            price = prices[0] if prices else None
    gbs = [g for g in extract_all_gb(text) if 6 <= g <= 30]
    gb = gbs[0] if gbs else None
    if gb is None:
        confidence = "low"
    return {"price": price, "plan_gb": gb, "confidence": confidence}


def parse_claro_controle(text: str) -> Optional[dict]:
    """Claro controle: no reliable primary keyword yet. Always low."""
    if not text or len(text) < 100:
        return None
    price = find_price_near_text(text, "Controle", window=500)
    if price is None:
        prices = [p for p in extract_all_prices(text) if 40 <= p <= 120]
        price = prices[0] if prices else None
    gbs = [g for g in extract_all_gb(text) if 10 <= g <= 60]
    gb = gbs[0] if gbs else None
    return {"price": price, "plan_gb": gb, "confidence": "low"}


def parse_claro_postpaid(text: str) -> Optional[dict]:
    """Claro post-paid: no reliable primary keyword yet. Always low."""
    if not text or len(text) < 100:
        return None
    price = find_price_near_text(text, "Pós", window=400)
    if price is None:
        prices = [p for p in extract_all_prices(text) if 80 <= p <= 250]
        price = prices[0] if prices else None
    gbs = [g for g in extract_all_gb(text) if 20 <= g <= 200]
    gb = gbs[0] if gbs else None
    return {"price": price, "plan_gb": gb, "confidence": "low"}


# ── Scraping job definitions ──────────────────────────────────────────────────
# Each job maps a segment/carrier to its URL, parser, and the JSON path inside
# plans.json so we know where to compare and update.
#
# price_range / gb_range: plausible bounds for what these plans cost today.
# Values outside these bounds are treated as parse errors (not real changes).
# Set range wide enough to catch real changes but tight enough to reject garbage.
# Rule of thumb: ±50% of current known value.

SCRAPE_JOBS = [
    {
        "segment": "prepaid", "carrier": "vivo",
        "url":     "https://vivo.com.br/para-voce/produtos-e-servicos/para-o-celular/pre-pago/vivo-pre",
        "parser":  parse_vivo_prepaid, "wait": 10,
        "price_range": (25, 55), "gb_range": (6, 20),
    },
    {
        "segment": "prepaid", "carrier": "tim",
        "url":     "https://www.tim.com.br/sp/para-voce/planos/pre-pago",
        "parser":  parse_tim_prepaid, "wait": 8,
        "price_range": (25, 55), "gb_range": (8, 20),
    },
    {
        "segment": "prepaid", "carrier": "claro",
        "url":     "https://www.claro.com.br/celular/planos-pre/prezao",
        "parser":  parse_claro_prepaid, "wait": 8,
        "price_range": (25, 55), "gb_range": (8, 20),
    },
    {
        "segment": "controle", "carrier": "vivo",
        "url":     "https://vivo.com.br/para-voce/produtos-e-servicos/para-o-celular/planos-controle",
        "parser":  parse_vivo_controle, "wait": 10,
        "price_range": (45, 100), "gb_range": (15, 40),
    },
    {
        "segment": "controle", "carrier": "tim",
        "url":     "https://www.tim.com.br/rj/para-voce/planos/controle",
        "parser":  parse_tim_controle, "wait": 8,
        "price_range": (45, 100), "gb_range": (10, 50),
    },
    {
        "segment": "controle", "carrier": "claro",
        "url":     "https://www.claro.com.br/celular/controle",
        "parser":  parse_claro_controle, "wait": 8,
        "price_range": (45, 100), "gb_range": (10, 50),
    },
    {
        "segment": "postpaid", "carrier": "vivo",
        "url":     "https://vivo.com.br/para-voce/produtos-e-servicos/para-o-celular/planos-pos-pago",
        "parser":  parse_vivo_postpaid, "wait": 10,
        "price_range": (100, 220), "gb_range": (30, 120),
    },
    {
        "segment": "postpaid", "carrier": "tim",
        "url":     "https://www.tim.com.br/rj/para-voce/planos/pos-pago/tim-black",
        "parser":  parse_tim_postpaid, "wait": 8,
        "price_range": (100, 250), "gb_range": (10, 80),
    },
    {
        "segment": "postpaid", "carrier": "claro",
        "url":     "https://www.claro.com.br/celular/pos",
        "parser":  parse_claro_postpaid, "wait": 8,
        "price_range": (80, 200), "gb_range": (20, 100),
    },
]


# ── Change detection ──────────────────────────────────────────────────────────

WATCHED_FIELDS = {
    "price":   ("price",),
    "plan_gb": ("gb", "plan_gb"),
}

FIELD_LABELS = {
    "price":   "Monthly Price (R$)",
    "plan_gb": "Base Plan GB",
}

SEGMENT_LABELS = {
    "prepaid":  "Pre-Paid (30d)",
    "controle": "Controle",
    "postpaid": "Post-Paid",
}

CARRIER_LABELS = {
    "vivo":  "Vivo",
    "tim":   "TIM",
    "claro": "Claro",
}


def get_nested(obj: dict, *keys: str) -> Any:
    """Safely navigate nested dict: get_nested(d, 'gb', 'plan_gb')."""
    for k in keys:
        if not isinstance(obj, dict):
            return None
        obj = obj.get(k)
    return obj


def set_nested(obj: dict, value: Any, *keys: str) -> None:
    """Set nested dict value: set_nested(d, 42, 'gb', 'plan_gb')."""
    for k in keys[:-1]:
        obj = obj.setdefault(k, {})
    obj[keys[-1]] = value


def find_plan(data: dict, segment: str, carrier: str) -> Optional[dict]:
    """Locate a plan object in the plans.json structure."""
    plans = data.get("segments", {}).get(segment, {}).get("plans", [])
    for p in plans:
        if p.get("carrier") == carrier:
            return p
    return None


def _in_range(value: Any, bounds: Optional[tuple]) -> bool:
    """Return True if value is within bounds (lo, hi), or if no bounds defined."""
    if bounds is None or value is None:
        return True
    lo, hi = bounds
    try:
        return lo <= float(value) <= hi
    except (TypeError, ValueError):
        return False


def detect_changes(
    data: dict,
    job: dict,
    scraped: dict,
) -> list[dict]:
    """
    Compare scraped values against stored plan. Returns list of change dicts.
    A change is only recorded when:
      - The scraped value is not None
      - The scraped value differs from the stored value
      - The scraped value falls within the job's plausible range (rejects garbage)
      - The plan is not marked scrape_status='manual'
    """
    segment = job["segment"]
    carrier = job["carrier"]
    plan = find_plan(data, segment, carrier)
    if not plan:
        log.warning(f"No plan found in JSON for {segment}/{carrier}")
        return []

    if plan.get("scrape_status") == "manual":
        log.info(f"Skipping {segment}/{carrier} — status is 'manual'")
        return []

    price_range = job.get("price_range")
    gb_range    = job.get("gb_range")

    changes = []
    for field_key, json_path in WATCHED_FIELDS.items():
        scraped_val = scraped.get(field_key)
        stored_val  = get_nested(plan, *json_path)

        if scraped_val is None:
            continue

        # Validate plausibility — discard values outside expected range
        bounds = price_range if field_key == "price" else gb_range
        if not _in_range(scraped_val, bounds):
            log.warning(
                f"Implausible scraped value for {carrier} {segment} {field_key}: "
                f"{scraped_val} (expected {bounds}) — discarding"
            )
            continue

        # Normalize types for comparison
        if field_key == "price":
            scraped_val = round(float(scraped_val), 2)
            stored_cmp  = round(float(stored_val), 2) if stored_val is not None else None
        else:
            scraped_val = int(scraped_val)
            stored_cmp  = int(stored_val) if stored_val is not None else None

        if stored_cmp is None or scraped_val != stored_cmp:
            changes.append({
                "segment":      segment,
                "carrier":      carrier,
                "plan_name":    plan.get("plan_name", ""),
                "field":        field_key,
                "field_label":  FIELD_LABELS[field_key],
                "old_value":    stored_val,
                "new_value":    scraped_val,
            })
            log.info(
                f"CHANGE detected: {carrier} {segment} — {field_key}: "
                f"{stored_val} → {scraped_val}"
            )

    return changes


def apply_changes(data: dict, job: dict, scraped: dict, changes: list[dict]) -> None:
    """Update plans.json in-memory with the new scraped values."""
    if not changes:
        return
    segment = job["segment"]
    carrier = job["carrier"]
    plan = find_plan(data, segment, carrier)
    if not plan:
        return

    for change in changes:
        field_key = change["field"]
        new_val   = change["new_value"]
        json_path = WATCHED_FIELDS[field_key]
        set_nested(plan, new_val, *json_path)
        log.info(f"Updated plans.json: {segment}/{carrier}/{'/'.join(json_path)} = {new_val}")

    plan["scrape_status"] = "ok"

    # Update history snapshot for today
    today = today_brt()
    history: list = data.setdefault("history", [])
    today_entry = next((h for h in history if h.get("date") == today), None)
    if today_entry is None:
        today_entry = {"date": today, "prepaid": {}, "controle": {}, "postpaid": {}}
        history.append(today_entry)

    seg_hist = today_entry.setdefault(segment, {})
    if "price" in scraped and scraped["price"] is not None:
        seg_hist.setdefault(carrier, {})["price"] = scraped["price"]
    if "plan_gb" in scraped and scraped["plan_gb"] is not None:
        seg_hist.setdefault(carrier, {})["plan_gb"] = scraped["plan_gb"]

    # Update meta
    data["meta"]["last_updated"] = today
    data["meta"]["collected_date_display"] = datetime.strptime(today, "%Y-%m-%d").strftime("%b %-d, %Y") if sys.platform != "win32" else datetime.strptime(today, "%Y-%m-%d").strftime("%b %d, %Y").replace(" 0", " ")


# ── XLSX ─────────────────────────────────────────────────────────────────────
# 4-tab workbook:
#   "Histórico de Mudanças" — one row per confirmed change
#   "Pré-Pago"              — one snapshot row per carrier per run
#   "Controle"              — idem
#   "Pós-Pago"              — idem

HEADER_FILL  = PatternFill("solid", fgColor="1A3A6B")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
CHANGE_FILL  = PatternFill("solid", fgColor="FFF3CD")   # amber for confirmed changes
WARN_FILL    = PatternFill("solid", fgColor="FFE0CC")   # orange for manual-needed
ALT_FILL     = PatternFill("solid", fgColor="F8F9FA")
BORDER_SIDE  = Side(style="thin", color="DEE2E6")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE,
)

SEGMENT_TABS = {
    "prepaid":  "Pré-Pago",
    "controle": "Controle",
    "postpaid": "Pós-Pago",
}

CHANGES_HEADERS = [
    "Data", "Hora (BRT)", "Empresa", "Segmento", "Plano",
    "Campo Alterado", "Valor Anterior", "Novo Valor", "Descrição",
]
CHANGES_WIDTHS = [12, 10, 10, 14, 28, 18, 16, 16, 40]

SNAPSHOT_HEADERS = [
    "Data", "Empresa", "Preço (R$)", "GB Base", "Bônus", "Status Coleta",
]
SNAPSHOT_WIDTHS = [12, 10, 12, 10, 30, 22]


def _write_headers(ws, headers: list[str], widths: list[int]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _load_or_create_workbook() -> openpyxl.Workbook:
    """Load existing 4-tab workbook or create it fresh."""
    if CHANGELOG_XLSX.exists():
        wb = openpyxl.load_workbook(CHANGELOG_XLSX)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Histórico de Mudanças"

    # Ensure all required tabs exist
    if "Histórico de Mudanças" not in wb.sheetnames:
        wb.create_sheet("Histórico de Mudanças", 0)
    for tab in SEGMENT_TABS.values():
        if tab not in wb.sheetnames:
            wb.create_sheet(tab)

    # Write headers on tabs that are brand-new (only row 1 exists or sheet is empty)
    ws_changes = wb["Histórico de Mudanças"]
    if ws_changes.max_row <= 1 and ws_changes.cell(1, 1).value is None:
        _write_headers(ws_changes, CHANGES_HEADERS, CHANGES_WIDTHS)

    for tab in SEGMENT_TABS.values():
        ws = wb[tab]
        if ws.max_row <= 1 and ws.cell(1, 1).value is None:
            _write_headers(ws, SNAPSHOT_HEADERS, SNAPSHOT_WIDTHS)

    return wb


def _get_bonus_summary(plan: dict) -> str:
    """Extract bonus GB description from a plan's legend (non-plan types)."""
    legend = plan.get("gb", {}).get("legend", [])
    parts = [
        item.get("amount", "")
        for item in legend
        if item.get("type") in ("bonus", "social", "port") and item.get("amount")
    ]
    return " / ".join(parts) if parts else "—"


def update_xlsx(
    snapshots: list[dict],
    changes: list[dict],
    dry_run: bool = False,
) -> None:
    """
    Update the 4-tab workbook:
    - Append confirmed changes to "Histórico de Mudanças"
    - Append one snapshot row per carrier to each segment tab
    """
    now      = now_brt()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M")

    if dry_run:
        log.info(f"[DRY RUN] Would write {len(snapshots)} snapshot(s) and "
                 f"{len(changes)} change(s) to changelog.xlsx")
        return

    CHANGELOG_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = _load_or_create_workbook()

    # ── Tab 1: Histórico de Mudanças ─────────────────────────────────────────
    ws_ch = wb["Histórico de Mudanças"]
    for ch in changes:
        row_idx = ws_ch.max_row + 1
        old_str = _format_value(ch["field"], ch["old_value"])
        new_str = _format_value(ch["field"], ch["new_value"])
        field_labels = {"price": "Preço", "plan_gb": "GB Base"}
        desc = f"{old_str} → {new_str}"
        values = [
            date_str, time_str,
            CARRIER_LABELS.get(ch["carrier"], ch["carrier"]),
            SEGMENT_LABELS.get(ch["segment"], ch["segment"]),
            ch.get("plan_name", ""),
            field_labels.get(ch["field"], ch["field"]),
            old_str, new_str, desc,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_ch.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="center")
            cell.fill      = CHANGE_FILL
            if col_idx in (7, 8):
                cell.font = Font(bold=True)

    # ── Tabs 2-4: per-segment snapshots ──────────────────────────────────────
    for snap in snapshots:
        tab = SEGMENT_TABS.get(snap["segment"])
        if not tab or tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        row_idx = ws.max_row + 1
        status  = snap["collection_status"]
        fill    = CHANGE_FILL if status == "confirmed_change" else (
                  WARN_FILL  if status == "manual_needed"    else
                  PatternFill())

        price_val = snap.get("price")
        price_str = f"{float(price_val):.2f}".replace(".", ",") if price_val is not None else "—"
        gb_str    = str(snap["plan_gb"]) if snap.get("plan_gb") is not None else "—"

        status_labels = {
            "confirmed":        "✅ Coletado",
            "confirmed_change": "✅ Atualizado",
            "manual_needed":    "⚠️ Verificar manualmente",
            "error":            "❌ Falha ao carregar",
        }
        values = [
            date_str,
            CARRIER_LABELS.get(snap["carrier"], snap["carrier"]),
            price_str,
            gb_str,
            snap.get("bonus", "—"),
            status_labels.get(status, status),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="center")
            cell.fill      = fill

    wb.save(CHANGELOG_XLSX)
    log.info(f"Workbook saved: {CHANGELOG_XLSX}")


# ── GitHub Issues notifier ────────────────────────────────────────────────────
# GitHub automatically sends an email to the repo owner when an issue is opened.
# No SMTP credentials, no external services — just the GITHUB_TOKEN that
# GitHub Actions provides automatically in every workflow run.

def _format_value(field: str, val: Any) -> str:
    if val is None:
        return "—"
    if field == "price":
        return f"R$ {float(val):.2f}".replace(".", ",")
    if field == "plan_gb":
        return f"{val} GB"
    return str(val)


def _pct_change(old: Any, new: Any) -> str:
    try:
        pct = (float(new) - float(old)) / float(old) * 100
        sign = "+" if pct >= 0 else ""
        return f"{sign}{pct:.1f}%"
    except Exception:
        return ""


def build_issue(
    snapshots: list[dict],
    changes: list[dict],
    scrape_errors: list[dict],
) -> tuple[str, str]:
    """
    Build GitHub Issue title and Markdown body.
    Returns (title, body_markdown).

    Shows one table per segment with all carriers, then a "what's missing" section.
    """
    today_str = now_brt().strftime("%Y-%m-%d")
    time_str  = now_brt().strftime("%H:%M BRT")
    n_changes = len(changes)
    missing   = [s for s in snapshots if s["collection_status"] in ("manual_needed", "error")]

    if n_changes > 0:
        title = f"⚠️ {n_changes} mudança(s) confirmada(s) — {today_str}"
    elif missing:
        title = f"🔍 {len(missing)} plano(s) precisam de verificação — {today_str}"
    else:
        title = f"✅ Verificação concluída — sem mudanças — {today_str}"

    lines = []
    lines.append(f"**BTG TMT — Telecom Pricing Monitor** · {today_str} às {time_str}\n")

    # ── Per-segment status tables ──────────────────────────────────────────────
    seg_order = [("prepaid", "Pré-Pago (30d)"), ("controle", "Controle"), ("postpaid", "Pós-Pago")]
    snap_index = {(s["segment"], s["carrier"]): s for s in snapshots}

    for seg_key, seg_label in seg_order:
        lines.append(f"## 📊 {seg_label}\n")
        lines.append("| Empresa | Preço | GB Base | Bônus | Status |")
        lines.append("|---------|-------|---------|-------|--------|")
        for carrier in ("vivo", "tim", "claro"):
            snap = snap_index.get((seg_key, carrier))
            if not snap:
                lines.append(f"| {CARRIER_LABELS.get(carrier, carrier)} | — | — | — | ❓ Sem dados |")
                continue
            price_str = (
                f"R$ {float(snap['price']):.2f}".replace(".", ",")
                if snap.get("price") is not None else "—"
            )
            gb_str    = f"{snap['plan_gb']} GB" if snap.get("plan_gb") is not None else "—"
            bonus_str = snap.get("bonus", "—")
            status    = snap["collection_status"]
            status_icon = {
                "confirmed":        "✅ Coletado",
                "confirmed_change": "✅ Atualizado automaticamente",
                "manual_needed":    "⚠️ Verificar — envie print",
                "error":            "❌ Falha ao carregar",
            }.get(status, status)
            lines.append(
                f"| **{CARRIER_LABELS.get(carrier, carrier)}** "
                f"| {price_str} | {gb_str} | {bonus_str} | {status_icon} |"
            )
        lines.append("")

    # ── Confirmed changes detail ───────────────────────────────────────────────
    if changes:
        lines.append(f"## ⚠️ {n_changes} Mudança(s) Confirmada(s)\n")
        lines.append("> Detectadas com alta confiança e aplicadas ao plans.json.\n")
        lines.append("| Empresa | Segmento | Plano | Campo | Antes | Depois | Δ |")
        lines.append("|---------|----------|-------|-------|-------|--------|---|")
        for ch in changes:
            old_fmt = _format_value(ch["field"], ch["old_value"])
            new_fmt = _format_value(ch["field"], ch["new_value"])
            pct     = _pct_change(ch["old_value"], ch["new_value"])
            field_labels = {"price": "Preço", "plan_gb": "GB Base"}
            lines.append(
                f"| **{CARRIER_LABELS.get(ch['carrier'], ch['carrier'])}** "
                f"| {SEGMENT_LABELS.get(ch['segment'], ch['segment'])} "
                f"| {ch.get('plan_name', '')} "
                f"| {field_labels.get(ch['field'], ch['field'])} "
                f"| ~~{old_fmt}~~ | **{new_fmt}** | {pct} |"
            )
        lines.append("")

    # ── What's missing ────────────────────────────────────────────────────────
    if missing:
        lines.append("## 🔍 Envie um print para estes planos\n")
        lines.append(
            "> Os valores abaixo **não foram atualizados**. "
            "Envie um screenshot do site da operadora no chat.\n"
        )
        for s in missing:
            reason = "site bloqueado" if s["collection_status"] == "error" else "marcador não encontrado na página"
            stored_price = (
                f"R$ {float(s['price']):.2f}".replace(".", ",")
                if s.get("price") is not None else "sem preço"
            )
            stored_gb = f"{s['plan_gb']} GB" if s.get("plan_gb") is not None else "sem GB"
            lines.append(
                f"- **{CARRIER_LABELS.get(s['carrier'], s['carrier'])} "
                f"{SEGMENT_LABELS.get(s['segment'], s['segment'])}** "
                f"— {reason}. Valor atual guardado: {stored_price} · {stored_gb}"
            )
        lines.append("")

    # ── Page errors ───────────────────────────────────────────────────────────
    if scrape_errors:
        lines.append(f"## ❌ Páginas que não carregaram ({len(scrape_errors)})\n")
        for e in scrape_errors:
            lines.append(
                f"- {CARRIER_LABELS.get(e['carrier'], e['carrier'])} "
                f"{SEGMENT_LABELS.get(e['segment'], e['segment'])}: {e.get('reason', '')}"
            )
        lines.append("")

    lines.append("---")
    lines.append("_Monitor automático BTG TMT · Pré-Pago · Controle · Pós-Pago · São Paulo DDD 11_")

    return title, "\n".join(lines)


def create_github_issue(title: str, body: str, dry_run: bool = False) -> bool:
    """
    Create a GitHub Issue via the REST API.
    GITHUB_TOKEN is provided automatically by Actions — no setup required.
    Returns True on success.
    """
    if dry_run:
        log.info(f"[DRY RUN] Would create GitHub Issue: {title}")
        return True

    if not GITHUB_TOKEN:
        log.warning("GITHUB_TOKEN not set — cannot create issue (OK for local runs)")
        return False

    if not GITHUB_REPO:
        log.warning("GITHUB_REPO not set — cannot create issue")
        return False

    # Label: create "pricing-monitor" label colour if it doesn't exist yet
    # (we don't fail if label API errors — issue creation is the critical path)
    label = "pricing-monitor"
    _ensure_label(label, "0d1f3c", "Automated pricing monitor alert")

    payload = json.dumps({
        "title":  title,
        "body":   body,
        "labels": [label],
    }).encode("utf-8")

    url = f"https://api.github.com/repos/{GITHUB_REPO}/issues"
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            "Authorization":        f"Bearer {GITHUB_TOKEN}",
            "Accept":               "application/vnd.github+json",
            "Content-Type":         "application/json",
            "X-GitHub-Api-Version": "2022-11-28",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
            issue_url = result.get("html_url", "")
            log.info(f"GitHub Issue created: {issue_url}")
            return True
    except urllib.error.HTTPError as e:
        body_text = e.read().decode("utf-8", errors="replace")
        log.error(f"GitHub API error {e.code}: {body_text}")
        return False
    except Exception as exc:
        log.error(f"Failed to create GitHub Issue: {exc}")
        return False


def _ensure_label(name: str, color: str, description: str) -> None:
    """Create the label in the repo if it doesn't already exist. Silent on failure."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return
    payload = json.dumps({"name": name, "color": color, "description": description}).encode()
    url = f"https://api.github.com/repos/{GITHUB_REPO}/labels"
    req = urllib.request.Request(
        url, data=payload,
        headers={
            "Authorization":        f"Bearer {GITHUB_TOKEN}",
            "Accept":               "application/vnd.github+json",
            "Content-Type":         "application/json",
            "X-GitHub-Api-Version": "2022-11-28",
        },
        method="POST",
    )
    try:
        urllib.request.urlopen(req, timeout=10)
    except Exception:
        pass  # label probably already exists — not a problem


# ── Main orchestration ────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="BTG TMT Telecom Pricing Monitor")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Scrape and detect changes but do NOT write files or create issues",
    )
    parser.add_argument(
        "--force-notify",
        action="store_true",
        help="Create a GitHub Issue even if no changes were detected (useful for testing)",
    )
    args = parser.parse_args()

    if args.dry_run:
        log.info("=== DRY RUN MODE — no files will be written, no issue will be created ===")

    # Load plans.json
    if not PLANS_JSON.exists():
        log.error(f"plans.json not found at {PLANS_JSON}")
        sys.exit(1)

    with open(PLANS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    all_changes:   list[dict] = []
    scrape_errors: list[dict] = []
    snapshots:     list[dict] = []   # one entry per plan, every run

    # Run all scrape jobs
    for job in SCRAPE_JOBS:
        seg     = job["segment"]
        carrier = job["carrier"]
        label   = f"{CARRIER_LABELS.get(carrier, carrier)} {SEGMENT_LABELS.get(seg, seg)}"
        plan    = find_plan(data, seg, carrier)

        log.info(f"--- Checking {label} ---")

        # Base snapshot from stored plans.json (always recorded, even on failure)
        snap: dict = {
            "segment":           seg,
            "carrier":           carrier,
            "plan_name":         plan.get("plan_name", "") if plan else "",
            "price":             plan.get("price") if plan else None,
            "plan_gb":           (plan.get("gb") or {}).get("plan_gb") if plan else None,
            "bonus":             _get_bonus_summary(plan) if plan else "—",
            "collection_status": "error",  # updated below if scraping succeeds
        }

        page_text = fetch_page_text(job["url"], wait_seconds=job.get("wait", 8))

        if page_text is None:
            log.warning(f"No page content for {label} — site blocked or timeout")
            scrape_errors.append({
                "carrier": carrier,
                "segment": seg,
                "reason":  "Site bloqueado ou timeout",
            })
            snapshots.append(snap)
            continue

        scraped = job["parser"](page_text)
        if scraped is None:
            log.warning(f"Parser returned None for {label}")
            scrape_errors.append({
                "carrier": carrier,
                "segment": seg,
                "reason":  "Página carregou mas conteúdo não foi parseado",
            })
            snapshots.append(snap)
            continue

        confidence = scraped.get("confidence", "low")
        log.info(
            f"{label}: scraped price={scraped.get('price')} "
            f"plan_gb={scraped.get('plan_gb')} confidence={confidence}"
        )

        if confidence == "low":
            snap["collection_status"] = "manual_needed"
            snapshots.append(snap)
            log.warning(f"Low confidence for {label} — stored values kept, flagged for manual check")
            continue

        # High confidence — check for changes
        job_changes = detect_changes(data, job, scraped)
        all_changes.extend(job_changes)

        if not args.dry_run and job_changes:
            apply_changes(data, job, scraped, job_changes)
            # Update snapshot price/gb to reflect what was just applied
            snap["price"]   = scraped.get("price", snap["price"])
            snap["plan_gb"] = scraped.get("plan_gb", snap["plan_gb"])
            snap["collection_status"] = "confirmed_change"
        else:
            snap["collection_status"] = "confirmed"

        snapshots.append(snap)

    # Write updated plans.json if there were confirmed changes
    if all_changes and not args.dry_run:
        with open(PLANS_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        log.info(f"plans.json saved with {len(all_changes)} change(s)")

    # Update XLSX (snapshots every run + confirmed changes)
    update_xlsx(snapshots, all_changes, dry_run=args.dry_run)

    # Create GitHub Issue — always, so you get an email every run
    title, body = build_issue(snapshots, all_changes, scrape_errors)
    create_github_issue(title, body, dry_run=args.dry_run)

    # Summary
    log.info(
        f"=== Done: {len(all_changes)} confirmed change(s), "
        f"{len([s for s in snapshots if s['collection_status'] == 'manual_needed'])} manual check(s) needed, "
        f"{len(scrape_errors)} page error(s) ==="
    )


if __name__ == "__main__":
    main()